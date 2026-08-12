from typing import Any
from datetime import date, datetime
import re

import io
import uuid
from datetime import date as date_type

from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.services.expense import ExpenseService
from app.schemas.expense import ExpenseCreate, ExpenseUpdate
from app.utils.dependencies import require_admin
from app.utils.response import success_response, paginated_response
from app.models.expense import Expense
from app.models.employee import Employee

router = APIRouter(prefix="/expenses", tags=["expenses"])


def _ex_to_dict(ex, emp: Employee | None) -> dict:
    if isinstance(ex, dict):
        d = ex.copy()
    else:
        d = {
            "id": ex.id,
            "employee_id": ex.employee_id,
            "product_name": getattr(ex, "product_name", None),
            "category": ex.category,
            "custom_category": getattr(ex, "custom_category", None),
            "amount": ex.amount,
            "description": ex.description or "",
            "expense_date": str(ex.expense_date) if ex.expense_date else "",
            "receipt_url": ex.receipt_url or "",
            "status": ex.status,
            "approved_by": ex.approved_by,
            "rejected_by": ex.rejected_by,
            "vendor": getattr(ex, "vendor", None),
            "department": getattr(ex, "department", None),
            "qty": getattr(ex, "qty", None),
            "created_at": ex.created_at.isoformat() if ex.created_at else None,
            "updated_at": ex.updated_at.isoformat() if ex.updated_at else None,
        }
    d["employee_name"] = f"{emp.first_name} {emp.last_name}" if emp else "Unknown"
    if "custom_category" not in d or d.get("custom_category") is None:
        try:
            if not isinstance(ex, dict):
                d["custom_category"] = getattr(ex, "custom_category", None)
        except Exception:
            d["custom_category"] = None
    # Category display name: if "other" and has custom_category, expose a display label
    d["category_display"] = (
        d.get("custom_category")
        if d.get("category") == "other" and d.get("custom_category")
        else (d.get("category") or "")
    )
    return d


_SORT_FIELDS = {"expense_date", "amount", "category", "product_name", "vendor", "status", "created_at"}


@router.get("")
def list_expenses(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
    category: str = Query(None),
    employee_id: str = Query(None),
    start_date: date = Query(None),
    end_date: date = Query(None),
    sort_by: str = Query("expense_date"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    skip = (page - 1) * per_page
    sort_by = sort_by if sort_by in _SORT_FIELDS else "expense_date"
    sort_order = "asc" if sort_order == "asc" else "desc"
    records, total = service.get_filtered(
        skip, per_page, category, employee_id, start_date, end_date, sort_by, sort_order
    )
    result = []
    for ex in records:
        eid = ex["employee_id"] if isinstance(ex, dict) else ex.employee_id
        emp = db.query(Employee).filter(Employee.id == eid).first()
        result.append(_ex_to_dict(ex, emp))
    return paginated_response(data=result, total=total, page=page, per_page=per_page)


class BulkDeleteRequest(BaseModel):
    ids: list[str]


@router.post("/bulk-delete")
def bulk_delete_expenses(
    data: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    deleted = service.bulk_delete(data.ids)
    return success_response(data={"deleted": deleted})


@router.get("/summary")
def get_expense_summary(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    return success_response(data=service.get_summary(start_date, end_date))


@router.get("/template")
def download_expense_template(current_user: Any = Depends(require_admin)):
    """Return a BLANK Excel file in the BIN OMOR TRADERS Expense Ledger layout
    so users know the exact format to fill in before uploading via Import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Details"

    bold_big = Font(bold=True, size=13)
    bold = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # A: Sl, B: Date, C: Vendor, D: Depertment, E: Item, F: Qty, G: Velue
    for col, w in [("A", 5), ("B", 14), ("C", 24), ("D", 22), ("E", 42), ("F", 8), ("G", 14)]:
        ws.column_dimensions[col].width = w

    # Row 1 — Brand header (merged A1:G1)
    ws.cell(row=1, column=1, value="BIN OMOR TRADERS — EXPENSE LEDGER")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).font = bold_big
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Row 2 — Address (merged A2:G2)
    ws.cell(row=2, column=1, value="House# 105/A, Road#1, Mohakhali D.O.H.S, Dhaka")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Row 3 — Date label + value
    ws.cell(row=3, column=1, value="Date")
    ws.cell(row=3, column=1).font = bold
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws.cell(row=3, column=2, value=date_type.today())
    ws.cell(row=3, column=2).number_format = "dd-mmm-yyyy"

    # Row 4 — Column headers
    for c, h in enumerate(["Sl", "Date", "Vendor", "Depertment", "Item", "Qty", "Velue"], 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 5-19 — Blank bordered rows for users to fill in.
    FIRST_DATA_ROW = 5
    LAST_DATA_ROW = 19
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin
        # Auto-fill serial numbers
        ws.cell(row=r, column=1, value=r - FIRST_DATA_ROW + 1)

    # Footer — Total Amount with auto-sum formula.
    total_row = LAST_DATA_ROW + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(row=total_row, column=1).font = bold
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    tot_cell = ws.cell(
        row=total_row,
        column=7,
        value=f"=SUM(G{FIRST_DATA_ROW}:G{LAST_DATA_ROW})",
    )
    tot_cell.font = bold
    tot_cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expense_template.xlsx"},
    )


# ─── Excel Import / Export helpers ───────────────────────────────

_EXP_HEADER_ALIASES = {
    "sl": "_sl", "si": "_sl", "serial": "_sl", "serialno": "_sl", "slno": "_sl",
    "date": "expense_date", "dt": "expense_date",
    "vendor": "vendor", "supplier": "vendor", "party": "vendor", "shop": "vendor",
    "department": "department", "depertment": "department", "dept": "department",
    "category": "department",
    "item": "product_name", "items": "product_name", "product": "product_name",
    "productname": "product_name",
    "description": "description", "details": "description",
    "particulars": "product_name", "particular": "product_name",
    "qty": "qty", "quantity": "qty", "qnty": "qty", "nos": "qty", "unit": "qty",
    "value": "amount", "velue": "amount", "amount": "amount",
    "price": "amount", "cost": "amount", "tk": "amount", "total": "amount",
}

# Maps department text (lower-cased, fuzzy) to the predefined Expense.category values.
_DEPT_TO_CATEGORY = {
    "furniture": "office_supplies",
    "electrical": "utilities",
    "hardware": "office_supplies",
    "internet": "communication",
    "stationery": "stationery",
    "travel": "travel",
    "transport": "transport",
    "fuel": "fuel",
    "food": "meals",
    "meal": "meals",
    "medical": "medical",
    "training": "training",
    "accommodation": "accommodation",
    "utility": "utilities",
    "utilities": "utilities",
    "communication": "communication",
    "office": "daily_office_needs",
}


def _exp_norm(s) -> str:
    """Normalise a header cell for fuzzy matching."""
    return re.sub(r"[^a-z0-9]", "", str(s or "").strip().lower())


def _exp_match_header(norm_val):
    """Return the field key for a normalised header string, or None."""
    if not norm_val:
        return None
    if norm_val in _EXP_HEADER_ALIASES:
        return _EXP_HEADER_ALIASES[norm_val]
    best, best_len = None, 0
    for alias_key, field_key in _EXP_HEADER_ALIASES.items():
        if alias_key in norm_val and len(alias_key) > best_len:
            best = field_key
            best_len = len(alias_key)
    return best


def _exp_to_date(val):
    """Coerce an Excel cell (datetime/date/str) into a datetime.date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date_type):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _dept_to_category(dept_text):
    """Map a free-text department to a (category, custom_category) pair."""
    if not dept_text:
        return ("other", None)
    norm = dept_text.strip().lower()
    for key, cat in _DEPT_TO_CATEGORY.items():
        if key in norm:
            return (cat, None)
    return ("other", dept_text.strip())


def _parse_expense_workbook(wb):
    """Parse an uploaded expense-ledger Excel into (title, address, ledger_date, [rows]).

    Layout expected (matching the Bin Omor Traders Expense Details sheet):
      Row 1  : Title (merged across columns)
      Row 2  : Address line (merged)
      Row 3  : "Date" label + a single date value for the whole ledger
      Row N  : Column header row (Sl | Date | Vendor | Depertment | Item | Qty | Velue)
      Row N+1: data rows until an empty row or a TOTAL row
      Side amounts in later columns (e.g. I/J) are imported as part of the row
      or appended as additional expense rows.
    """
    ws = wb.active

    # 1) Find the column-header row.
    header_row_idx = None
    col_map: dict[int, str] = {}
    for r in range(1, min(ws.max_row, 50) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)]
        matched = {}
        for c, val in enumerate(row_vals, 1):
            key = _exp_match_header(_exp_norm(val))
            if key and key not in matched.values():
                matched[c] = key
        vals_set = set(matched.values())
        if "amount" in vals_set and (
            "product_name" in vals_set or "description" in vals_set or "vendor" in vals_set
        ):
            header_row_idx = r
            col_map = matched
            break

    if header_row_idx is None:
        raise HTTPException(
            400,
            "Could not find a column-header row (Sl/Date/Vendor/Item/Velue). "
            "Please ensure your sheet has a header row with these columns.",
        )

    # Pre-header rows: row1 -> title, row2 -> address, any "Date" cell -> ledger_date.
    title = None
    address_line = None
    ledger_date = None
    for r in range(1, header_row_idx):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)]
        if ledger_date is None:
            for c, val in enumerate(row_vals, 1):
                if _exp_norm(val) == "date":
                    for c2 in range(c + 1, len(row_vals) + 1):
                        dv = ws.cell(row=r, column=c2).value
                        if dv not in (None, ""):
                            ledger_date = _exp_to_date(dv)
                            break
                    break
        cell_val = row_vals[0] if row_vals else None
        if cell_val is None or str(cell_val).strip() == "" or _exp_norm(cell_val) == "date":
            continue
        txt = str(cell_val).strip()
        if title is None:
            title = txt
        elif address_line is None:
            address_line = txt

    # 2) Read data rows after the header row.
    rows = []
    side_pairs = []
    main_cols = set(col_map.keys())
    r = header_row_idx + 1
    while r <= ws.max_row:
        row_vals = {c: ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)}
        if all((v is None or str(v).strip() == "") for v in row_vals.values()):
            r += 1
            continue
        joined = " ".join(str(v) for v in row_vals.values() if v is not None).lower()
        if "total" in joined:
            r += 1
            continue

        # Side table: cells in non-main columns forming (label, number) pairs.
        for c in range(1, min(ws.max_column, 15) + 1):
            if c in main_cols:
                continue
            val = row_vals.get(c)
            if isinstance(val, (int, float)) and val > 0:
                label_col = c - 1
                lbl = row_vals.get(label_col)
                if lbl and isinstance(lbl, str) and lbl.strip():
                    side_pairs.append((str(lbl).strip(), float(val)))

        rec = {
            "vendor": None, "department": None, "product_name": None,
            "description": None, "qty": None, "amount": None,
            "expense_date": ledger_date,
        }

        for c, key in col_map.items():
            if key == "_sl":
                continue
            val = row_vals.get(c)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            if key == "qty":
                rec["qty"] = str(val).strip()
            elif key == "amount":
                try:
                    rec["amount"] = float(val)
                except (ValueError, TypeError):
                    rec["amount"] = None
            elif key == "expense_date":
                rec["expense_date"] = _exp_to_date(val)
            else:
                rec[key] = str(val).strip()

        # If amount cell empty, scan for stray numeric values (split-amount quirk).
        # ONLY do this when the row has identifying data (vendor/item/dept) —
        # otherwise we'd pick up continuation/breakdown rows that are just
        # sub-amounts of the previous main expense.
        if rec["amount"] is None and (rec["product_name"] or rec["vendor"] or rec["department"]):
            max_main = max(main_cols) if main_cols else 1
            for c in range(max_main + 1, min(ws.max_column, 15) + 1):
                val = row_vals.get(c)
                if isinstance(val, (int, float)) and val > 0:
                    rec["amount"] = float(val)
                    break

        # Only add rows that have BOTH an amount AND some descriptive data,
        # or rows with descriptive data even without amount (rare).
        # This skips bare continuation/breakdown rows (serial + side amount only).
        if rec["amount"] is not None and (rec["product_name"] or rec["vendor"] or rec["department"]):
            rows.append(rec)
        elif rec["product_name"] or rec["vendor"]:
            rows.append(rec)
        r += 1

    # 3) Append side-table misc costs as additional expense rows.
    for label, amt in side_pairs:
        rows.append({
            "vendor": None, "department": None,
            "product_name": label, "description": None,
            "qty": None, "amount": amt,
            "expense_date": ledger_date,
        })

    if not title:
        title = "Imported Expense Ledger"

    return title, address_line, ledger_date, rows


@router.post("/import-excel")
async def import_expense_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    """Upload an Excel expense ledger (.xlsx) and bulk-import all rows as
    standalone Expense records."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload a .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read Excel file: {exc}")

    try:
        title, address_line, ledger_date, rows = _parse_expense_workbook(wb)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Could not parse the Excel layout: {exc}")

    try:
        created = 0
        for rec in rows:
            amt = rec.get("amount")
            if amt is None:
                continue
            dept = rec.get("department")
            category, custom_category = _dept_to_category(dept)
            desc_parts = []
            if rec.get("vendor"):
                desc_parts.append(f"Vendor: {rec['vendor']}")
            if rec.get("product_name"):
                desc_parts.append(rec["product_name"])
            if rec.get("description"):
                desc_parts.append(rec["description"])
            description = " | ".join(desc_parts) if desc_parts else (title or "Imported expense")

            exp = Expense(
                id=str(uuid.uuid4()),
                employee_id=None,
                product_name=rec.get("product_name") or None,
                category=category,
                custom_category=custom_category,
                amount=amt,
                description=description,
                expense_date=rec.get("expense_date") or date_type.today(),
                status="pending",
                vendor=rec.get("vendor"),
                department=dept,
                qty=rec.get("qty"),
            )
            db.add(exp)
            created += 1

        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(500, f"Failed to save expenses: {exc}")

    return {
        "success": True,
        "imported_rows": created,
        "title": title,
        "message": f"{created} expense row(s) imported from '{title}'",
    }


def _build_expense_ledger_excel(expenses, title="BIN OMOR TRADERS - Expense Ledger",
                                address="House# 105/A, Road#1, Mohakhali D.O.H.S, Dhaka",
                                ledger_date=None) -> bytes:
    """Build a Bin Omor Traders-style Excel ledger from a list of Expense records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Details"

    bold_big = Font(bold=True, size=13)
    bold = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(vertical="top", wrap_text=True)
    money_fmt = "#,##0.00"

    for col, w in [("A", 5), ("B", 14), ("C", 24), ("D", 22), ("E", 42), ("F", 8), ("G", 14)]:
        ws.column_dimensions[col].width = w

    row = 1
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).font = bold_big
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.cell(row=row, column=1, value=address)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.cell(row=row, column=1, value="Date")
    ws.cell(row=row, column=1).font = bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    if ledger_date:
        ws.cell(row=row, column=2, value=ledger_date)
        ws.cell(row=row, column=2).number_format = "dd-mmm-yyyy"
    row += 1

    for c, h in enumerate(["Sl", "Date", "Vendor", "Depertment", "Item", "Qty", "Velue"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    total = 0.0
    sl = 1
    for ex in expenses:
        ws.cell(row=row, column=1, value=sl)
        ed = getattr(ex, "expense_date", None) if not isinstance(ex, dict) else ex.get("expense_date")
        if ed:
            ws.cell(row=row, column=2, value=ed)
            ws.cell(row=row, column=2).number_format = "dd-mmm-yyyy"
        vendor = getattr(ex, "vendor", None) if not isinstance(ex, dict) else ex.get("vendor")
        ws.cell(row=row, column=3, value=vendor or "")
        dept = getattr(ex, "department", None) if not isinstance(ex, dict) else ex.get("department")
        ws.cell(row=row, column=4, value=dept or "")
        pn = getattr(ex, "product_name", None) if not isinstance(ex, dict) else ex.get("product_name")
        desc = getattr(ex, "description", None) if not isinstance(ex, dict) else ex.get("description")
        ws.cell(row=row, column=5, value=pn or desc or "")
        qty = getattr(ex, "qty", None) if not isinstance(ex, dict) else ex.get("qty")
        ws.cell(row=row, column=6, value=qty or "")
        amt = getattr(ex, "amount", 0) if not isinstance(ex, dict) else ex.get("amount", 0)
        amt_cell = ws.cell(row=row, column=7, value=amt)
        amt_cell.number_format = money_fmt
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = thin
            ws.cell(row=row, column=c).alignment = wrap
        total += float(amt or 0)
        sl += 1
        row += 1

    if expenses:
        row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    tot_cell = ws.cell(row=row, column=7, value=total)
    tot_cell.font = bold
    tot_cell.number_format = money_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/download-excel")
def download_expense_excel(
    category: str = Query(None),
    employee_id: str = Query(None),
    start_date: date = Query(None),
    end_date: date = Query(None),
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    """Export all expenses (optionally filtered) as a BIN OMOR TRADERS-style Excel ledger."""
    service = ExpenseService(db)
    records, _total = service.get_filtered(0, 10000, category, employee_id, start_date, end_date)
    safe_title = "Expense_Ledger"
    xls_bytes = _build_expense_ledger_excel(
        records,
        title=f"BIN OMOR TRADERS - Expense Ledger ({date_type.today().isoformat()})",
    )
    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_title}_{date_type.today().isoformat()}.xlsx"},
    )


@router.get("/{expense_id}")
def get_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    ex = service.get_by_id(expense_id)
    emp = db.query(Employee).filter(Employee.id == (ex.employee_id if not isinstance(ex, dict) else ex["employee_id"])).first()
    return success_response(data=_ex_to_dict(ex, emp))


@router.post("")
def create_expense(
    data: ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    payload = data.model_dump(exclude_unset=False)
    if payload.get("category") == "other" and payload.get("custom_category"):
        pass
    if not payload.get("amount") or payload["amount"] <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    if not payload.get("expense_date"):
        raise HTTPException(status_code=400, detail="Expense date is required")
    ex = service.create(payload)
    emp = db.query(Employee).filter(Employee.id == (ex.employee_id if not isinstance(ex, dict) else ex["employee_id"])).first() if (ex.employee_id if not isinstance(ex, dict) else ex.get("employee_id")) else None
    return success_response(data=_ex_to_dict(ex, emp))


@router.put("/{expense_id}")
def update_expense(
    expense_id: str,
    data: ExpenseUpdate,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    update_data = data.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    ex = service.update(expense_id, update_data)
    emp = db.query(Employee).filter(Employee.id == (ex.employee_id if not isinstance(ex, dict) else ex["employee_id"])).first() if (ex.employee_id if not isinstance(ex, dict) else ex.get("employee_id")) else None
    return success_response(data=_ex_to_dict(ex, emp))


@router.delete("/{expense_id}")
def delete_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    service.delete(expense_id)
    return success_response(data=None)


@router.post("/{expense_id}/approve")
def approve_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    ex = service.approve(expense_id, current_user.id)
    emp = db.query(Employee).filter(Employee.id == (ex.employee_id if not isinstance(ex, dict) else ex["employee_id"])).first()
    return success_response(data=_ex_to_dict(ex, emp))


@router.post("/{expense_id}/reject")
def reject_expense(
    expense_id: str,
    db: Session = Depends(get_db),
    current_user: Any = Depends(require_admin),
):
    service = ExpenseService(db)
    ex = service.reject(expense_id, current_user.id)
    emp = db.query(Employee).filter(Employee.id == (ex.employee_id if not isinstance(ex, dict) else ex["employee_id"])).first()
    return success_response(data=_ex_to_dict(ex, emp))
