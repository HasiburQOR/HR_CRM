import io
import os
import uuid
from datetime import datetime, timezone, date as date_type

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.requisition import Requisition, RequisitionExpense
from app.utils.dependencies import get_current_user, require_admin

router = APIRouter(prefix="/requisitions", tags=["Requisitions"])

UPLOAD_DIR = "/app/uploads/receipts"
os.makedirs(UPLOAD_DIR, exist_ok=True)

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}


def _req_to_dict(r: Requisition) -> dict:
    return {
        "id": r.id,
        "title": r.title,
        "status": r.status,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "closed_at": r.closed_at.isoformat() if r.closed_at else None,
        "duration_days": r.duration_days,
        "address": getattr(r, "address", None),
        "period": getattr(r, "period", None),
        "ledger_date": (getattr(r, "ledger_date", None).isoformat() if getattr(r, "ledger_date", None) else None),
    }


def _exp_to_dict(e: RequisitionExpense) -> dict:
    return {
        "id": e.id,
        "requisition_id": e.requisition_id,
        "expense_date": e.expense_date.isoformat() if e.expense_date else None,
        "notes": e.notes,
        "amount": e.amount,
        "receipt_url": e.receipt_url,
        "vendor": getattr(e, "vendor", None),
        "department": getattr(e, "department", None),
        "qty": getattr(e, "qty", None),
        "status": getattr(e, "status", None) or "pending",
        "approved_by": getattr(e, "approved_by", None),
        "rejected_by": getattr(e, "rejected_by", None),
        "created_at": e.created_at.isoformat() if e.created_at else None,
        "updated_at": (getattr(e, "updated_at", None).isoformat() if getattr(e, "updated_at", None) else None),
    }


# ─── CRUD ────────────────────────────────────────────────────────

@router.post("")
def create_requisition(
    title: str = Form(...),
    address: str = Form(None),
    period: str = Form(None),
    ledger_date: date_type = Form(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    req = Requisition(
        id=str(uuid.uuid4()),
        title=title,
        status="open",
        address=(address or None),
        period=(period or None),
        ledger_date=ledger_date,
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    return {"success": True, "data": _req_to_dict(req), "message": "Requisition created"}


@router.get("")
def list_requisitions(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    rows = db.query(Requisition).order_by(Requisition.created_at.desc()).all()
    return {"success": True, "data": [_req_to_dict(r) for r in rows]}


@router.get("/template")
def download_template(current_user=Depends(get_current_user)):
    """Return a BLANK Excel file in the Bin Omor layout so users know the exact
    format to fill in before uploading via the Import button."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    bold_big = Font(bold=True, size=13)
    bold = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, w in [("A", 6), ("B", 40), ("C", 12), ("D", 16)]:
        ws.column_dimensions[col].width = w

    # Row 1 — Brand header (merged A1:D1)
    ws.cell(row=1, column=1, value="BIN OMOR TRADERS — Requisition List")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1).font = bold_big
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Row 2 — Address (merged A2:D2)
    ws.cell(row=2, column=1, value="House# 105/A, Road#1, Mohakhali D.O.H.S, Dhaka")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Row 3 — Date label + value (the day the template was created)
    ws.cell(row=3, column=1, value="Date")
    ws.cell(row=3, column=1).font = bold
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
    ws.cell(row=3, column=2, value=date_type.today())
    ws.cell(row=3, column=2).number_format = "dd-mmm-yyyy"

    # Row 4 — Column headers
    for c, h in enumerate(["Sl.", "Item", "Qty", "Value (approx)"], 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 5-14 — Blank bordered rows for users to fill in.
    FIRST_DATA_ROW = 5
    LAST_DATA_ROW = 14
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = thin

    # Footer — Approx Total Amount with auto-sum formula.
    total_row = LAST_DATA_ROW + 1
    ws.cell(row=total_row, column=1, value="Approx Total Amount")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=1).font = bold
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    tot_cell = ws.cell(
        row=total_row,
        column=4,
        value=f"=SUM(D{FIRST_DATA_ROW}:D{LAST_DATA_ROW})",
    )
    tot_cell.font = bold
    tot_cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=requisition_template.xlsx"},
    )


@router.get("/{req_id}")
def get_requisition(req_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    data = _req_to_dict(req)
    data["expenses"] = [_exp_to_dict(e) for e in req.expenses]
    return {"success": True, "data": data}


# ─── Edit ────────────────────────────────────────────────────────

@router.put("/{req_id}")
def edit_requisition(
    req_id: str,
    title: str = Form(...),
    address: str = Form(None),
    period: str = Form(None),
    ledger_date: date_type = Form(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    title = (title or "").strip()
    if not title:
        raise HTTPException(400, "Title is required")
    req.title = title
    req.address = address or None
    req.period = period or None
    req.ledger_date = ledger_date
    db.commit()
    db.refresh(req)
    data = _req_to_dict(req)
    data["expenses"] = [_exp_to_dict(e) for e in req.expenses]
    return {"success": True, "data": data, "message": "Requisition updated"}


# ─── Close ───────────────────────────────────────────────────────

@router.put("/{req_id}/close")
def close_requisition(req_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.status == "closed":
        raise HTTPException(400, "Requisition is already closed")
    now = datetime.now(timezone.utc)
    req.closed_at = now
    req.status = "closed"
    if req.created_at:
        delta = now - req.created_at
        req.duration_days = max(1, delta.days)
    db.commit()
    db.refresh(req)
    data = _req_to_dict(req)
    data["expenses"] = [_exp_to_dict(e) for e in req.expenses]
    return {"success": True, "data": data, "message": "Requisition closed"}


# ─── Reopen ──────────────────────────────────────────────────────

@router.put("/{req_id}/reopen")
def reopen_requisition(req_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.status == "open":
        raise HTTPException(400, "Requisition is already open")
    req.status = "open"
    req.closed_at = None
    req.duration_days = None
    db.commit()
    db.refresh(req)
    data = _req_to_dict(req)
    data["expenses"] = [_exp_to_dict(e) for e in req.expenses]
    return {"success": True, "data": data, "message": "Requisition reopened"}


# ─── Expenses (multipart) ────────────────────────────────────────

@router.post("/{req_id}/expenses")
async def add_expense(
    req_id: str,
    note: str = Form(None),
    amount: float = Form(...),
    expense_date: date_type = Form(None),
    vendor: str = Form(None),
    department: str = Form(None),
    qty: str = Form(None),
    receipt: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.status == "closed":
        raise HTTPException(400, "Cannot add expenses to a closed requisition")

    receipt_url = None
    if receipt:
        ext = os.path.splitext(receipt.filename or "file")[1] or ".bin"
        filename = f"{uuid.uuid4()}{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)
        content = await receipt.read()
        with open(filepath, "wb") as f:
            f.write(content)
        receipt_url = f"/uploads/receipts/{filename}"

    exp = RequisitionExpense(
        id=str(uuid.uuid4()),
        requisition_id=req_id,
        expense_date=expense_date or date_type.today(),
        notes=note,
        amount=amount,
        receipt_url=receipt_url,
        vendor=(vendor or None),
        department=(department or None),
        qty=qty,
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return {"success": True, "data": _exp_to_dict(exp), "message": "Expense added"}


# ─── Expense Approval (admin only) ───────────────────────────────

@router.post("/{req_id}/expenses/{exp_id}/approve")
def approve_expense(
    req_id: str,
    exp_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    exp = (
        db.query(RequisitionExpense)
        .filter(RequisitionExpense.id == exp_id, RequisitionExpense.requisition_id == req_id)
        .first()
    )
    if not exp:
        raise HTTPException(404, "Expense not found")
    exp.status = "approved"
    exp.approved_by = str(current_user.id)
    exp.rejected_by = None
    db.commit()
    db.refresh(exp)
    return {"success": True, "data": _exp_to_dict(exp), "message": "Expense approved"}


@router.post("/{req_id}/expenses/{exp_id}/reject")
def reject_expense(
    req_id: str,
    exp_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    exp = (
        db.query(RequisitionExpense)
        .filter(RequisitionExpense.id == exp_id, RequisitionExpense.requisition_id == req_id)
        .first()
    )
    if not exp:
        raise HTTPException(404, "Expense not found")
    exp.status = "rejected"
    exp.rejected_by = str(current_user.id)
    exp.approved_by = None
    db.commit()
    db.refresh(exp)
    return {"success": True, "data": _exp_to_dict(exp), "message": "Expense rejected"}


# ─── Serve Receipt Files ─────────────────────────────────────────

@router.get("/receipts/{filename}")
def serve_receipt(filename: str, current_user=Depends(get_current_user)):
    filepath = os.path.join(UPLOAD_DIR, filename)
    if not os.path.isfile(filepath):
        raise HTTPException(404, "Receipt not found")
    return FileResponse(filepath)


# ─── Download Reports (Excel with Receipts) ──────────────────────

def _add_image_to_sheet(ws, row, col, filepath, max_w_px=200, max_h_px=130):
    """Embed a receipt image into the worksheet, scaling to fit and adjusting row/col size."""
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XlImage
        pil_img = PILImage.open(filepath)
        orig_w, orig_h = pil_img.size

        # Scale down if larger than max, preserving aspect ratio
        ratio = min(max_w_px / orig_w, max_h_px / orig_h, 1.0)
        target_w = int(orig_w * ratio)
        target_h = int(orig_h * ratio)

        # Fallback to small defaults if image is tiny
        if target_w < 10:
            target_w = 80
        if target_h < 10:
            target_h = 60

        img = XlImage(filepath)
        img.width = target_w
        img.height = target_h

        # Adjust row height (Excel row height is in points; 1 point ≈ 1.33 px)
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, target_h * 0.75 + 6)

        # Adjust column width (Excel col width is in char units; ~7 px per unit)
        col_letter = get_column_letter(col)
        needed_char_w = target_w / 7
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 8, needed_char_w)

        cell_ref = f"{col_letter}{row}"
        ws.add_image(img, cell_ref)
    except Exception:
        ws.cell(row=row, column=col, value="(image error)")


def _build_excel(requisitions: list, db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisitions"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = '#,##0.00'

    # Headers
    headers = [
        "Requisition Title", "Status", "Created At", "Closed At",
        "Duration (days)", "Expense Date", "Vendor", "Department", "Qty",
        "Note", "Amount (৳)", "Receipt", "Approval Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Column widths
    widths = [28, 10, 20, 20, 12, 14, 20, 18, 8, 30, 14, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for req in requisitions:
        exps = (
            db.query(RequisitionExpense)
            .filter(RequisitionExpense.requisition_id == req.id)
            .order_by(RequisitionExpense.created_at)
            .all()
        )
        if not exps:
            for c, val in enumerate([
                req.title, req.status,
                req.created_at.strftime("%Y-%m-%d %H:%M") if req.created_at else "",
                req.closed_at.strftime("%Y-%m-%d %H:%M") if req.closed_at else "",
                req.duration_days or "", "", "", "", "", "", "", "", "",
            ], 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = thin_border
            row += 1
        else:
            for idx, e in enumerate(exps):
                vals = [
                    req.title if idx == 0 else "",
                    req.status if idx == 0 else "",
                    req.created_at.strftime("%Y-%m-%d %H:%M") if idx == 0 and req.created_at else "",
                    req.closed_at.strftime("%Y-%m-%d %H:%M") if idx == 0 and req.closed_at else "",
                    req.duration_days if idx == 0 else "",
                    e.expense_date.isoformat() if e.expense_date else "",
                    getattr(e, "vendor", None) or "",
                    getattr(e, "department", None) or "",
                    getattr(e, "qty", None) if getattr(e, "qty", None) is not None else "",
                    e.notes or "",
                    e.amount,
                    "",
                    (getattr(e, "status", None) or "pending"),
                ]
                for c, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=val)
                    cell.border = thin_border
                    if c == 11 and isinstance(val, (int, float)):
                        cell.number_format = money_fmt

                # Embed receipt image if available
                if e.receipt_url:
                    fname = os.path.basename(e.receipt_url)
                    fpath = os.path.join(UPLOAD_DIR, fname)
                    if os.path.isfile(fpath) and os.path.splitext(fpath)[1].lower() in IMAGE_EXTS:
                        _add_image_to_sheet(ws, row, 12, fpath)
                    else:
                        ws.cell(row=row, column=12, value="(non-image receipt)")

                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_ledger_excel(req: Requisition, db: Session) -> bytes:
    """Build a single-requisition Excel that reproduces the Bin Omor Traders
    Requisition List layout:

      Row 1 (merged A1:D1): Title
      Row 2 (merged A2:D2): Address
      Row 3:  "Date" | <ledger_date>
      Row 4:  Sl. | Item | Qty | Value (approx)
      Row 5+: data rows
      Last:   "Approx Total Amount" | <sum>  (label merged A:C)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = req.title[:31] if req.title else "Requisition"

    bold = Font(bold=True, size=11)
    bold_big = Font(bold=True, size=13)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)
    money_fmt = '#,##0.00'

    for col, w in [("A", 6), ("B", 40), ("C", 12), ("D", 16)]:
        ws.column_dimensions[col].width = w

    row = 1
    # Row 1 — Title (merged A1:D1)
    ws.cell(row=row, column=1, value=req.title or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).font = bold_big
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Row 2 — Address (merged A2:D2)
    ws.cell(row=row, column=1, value=getattr(req, "address", None) or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Row 3 — Date label + value
    ws.cell(row=row, column=1, value="Date")
    ws.cell(row=row, column=1).font = bold
    ld = getattr(req, "ledger_date", None)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    if ld:
        ws.cell(row=row, column=2, value=ld)
        ws.cell(row=row, column=2).number_format = "dd-mmm-yyyy"
    row += 1

    # Row 4 — Column headers
    for c, h in enumerate(["Sl.", "Item", "Qty", "Value (approx)"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    # Data rows
    exps = (
        db.query(RequisitionExpense)
        .filter(RequisitionExpense.requisition_id == req.id)
        .order_by(RequisitionExpense.created_at)
        .all()
    )
    total = 0.0
    sl = 1
    for e in exps:
        ws.cell(row=row, column=1, value=sl)
        ws.cell(row=row, column=2, value=e.notes or "")
        ws.cell(row=row, column=3, value=getattr(e, "qty", None) or "")
        amt_cell = ws.cell(row=row, column=4, value=e.amount)
        amt_cell.number_format = money_fmt
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = thin
            ws.cell(row=row, column=c).alignment = wrap
        total += float(e.amount or 0)
        sl += 1
        row += 1

    # Total row (label merged A:C, amount in D)
    if exps:
        row += 1  # blank gap row like the original
    ws.cell(row=row, column=1, value="Approx Total Amount")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    tot_cell = ws.cell(row=row, column=4, value=total)
    tot_cell.font = bold
    tot_cell.number_format = money_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{req_id}/download")
def download_single_requisition(
    req_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Export a single requisition as a BIN OMOR TRADERS-style Excel ledger."""
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    xls_bytes = _build_ledger_excel(req, db)
    # Sanitise the filename: replace spaces/slashes, then strip any non-ASCII
    # chars (e.g. the em-dash \u2014 in "BIN OMOR TRADERS — …") because HTTP
    # headers are latin-1 and Starlette will crash on them.
    import re as _re
    safe_title = _re.sub(r"\s+", "_", (req.title or "requisition")).replace("/", "-")
    safe_title = safe_title.encode("ascii", "ignore").decode("ascii")[:50].strip("_") or "requisition"
    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_title}_{date_type.today().isoformat()}.xlsx"},
    )


class BulkDownloadRequest(BaseModel):
    ids: list[str]


@router.post("/download-bulk")
def download_bulk(
    payload: BulkDownloadRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    reqs = db.query(Requisition).filter(Requisition.id.in_(payload.ids)).all()
    if not reqs:
        raise HTTPException(404, "No requisitions found")
    xls_bytes = _build_excel(reqs, db)
    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bulk_requisitions_{date_type.today().isoformat()}.xlsx"},
    )


# ─── Import from Excel (auto-entry) ───────────────────────────────

# Header keywords → field key. Matched case-insensitively (handles typos like
# "Velue") by normalising: strip, lower, remove non-alphanumerics.
_HEADER_ALIASES = {
    "sl": "_sl", "si": "_sl", "serial": "_sl", "serialno": "_sl", "slno": "_sl",
    "item": "notes", "items": "notes", "description": "notes",
    "particulars": "notes", "details": "notes", "particular": "notes",
    "qty": "qty", "quantity": "qty", "qnty": "qty", "nos": "qty", "unit": "qty",
    "value": "amount", "velue": "amount", "amount": "amount", "price": "amount",
    "total": "amount", "cost": "amount", "tk": "amount",
    # optional columns (used only if present)
    "vendor": "vendor", "supplier": "vendor", "party": "vendor", "shop": "vendor",
    "department": "department", "depertment": "department",
    "dept": "department", "category": "department",
}


def _norm(s) -> str:
    """Normalise a header cell for fuzzy matching."""
    import re
    return re.sub(r"[^a-z0-9]", "", str(s or "").strip().lower())


def _match_header(norm_val):
    """Return the field key for a normalised header string, or None.

    Exact alias match first, then falls back to substring matching so headers
    like 'Value (approx)' -> 'valueapprox' still map to 'amount'.
    """
    if not norm_val:
        return None
    if norm_val in _HEADER_ALIASES:
        return _HEADER_ALIASES[norm_val]
    best, best_len = None, 0
    for alias_key, field_key in _HEADER_ALIASES.items():
        if alias_key in norm_val and len(alias_key) > best_len:
            best = field_key
            best_len = len(alias_key)
    return best


def _to_date(val):
    """Coerce an Excel cell (datetime/date/str) into a datetime.date."""
    from datetime import datetime as _dt
    if val is None:
        return None
    # IMPORTANT: datetime is a subclass of date, so check datetime FIRST.
    # Otherwise isinstance(val, date) swallows datetimes and returns them
    # unconverted, which breaks Date columns in some DBs.
    if isinstance(val, _dt):
        return val.date()
    if isinstance(val, date_type):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_import_workbook(wb):
    """Parse an uploaded Excel ledger into (title, address, period, ledger_date, [rows]).

    Layout expected (matching the Bin Omor Traders Requisition List):
      Row 1  : Title (merged across columns)
      Row 2  : Address line (merged)
      Row 3  : "Date" label + a single date value for the whole requisition
      Row N  : Column header row (Sl | Item | Qty | Value)
      Row N+1: data rows until an empty row or a TOTAL row
      Optionally a side table in later columns (e.g. E/F): label + amount pairs
      that are imported as additional expense rows.
    """
    ws = wb.active

    # 1) Find the column-header row.
    header_row_idx = None
    col_map: dict[int, str] = {}
    for r in range(1, min(ws.max_row, 50) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)]
        matched = {}
        for c, val in enumerate(row_vals, 1):
            key = _match_header(_norm(val))
            if key and key not in matched.values():
                matched[c] = key
        # Require at least item + amount columns to recognise a header row.
        vals_set = set(matched.values())
        if "notes" in vals_set and "amount" in vals_set:
            header_row_idx = r
            col_map = matched
            break

    if header_row_idx is None:
        raise HTTPException(400, "Could not find a column-header row (Sl/Item/Value). "
                                 "Please ensure your sheet has a header row with these columns.")

    # Pre-header rows: row1 -> title, row2 -> address, any "Date" cell -> ledger_date.
    title = None
    address_line = None
    ledger_date = None
    for r in range(1, header_row_idx):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)]
        # Detect a "Date: <value>" pair anywhere in the pre-header rows.
        if ledger_date is None:
            for c, val in enumerate(row_vals, 1):
                if _norm(val) == "date":
                    for c2 in range(c + 1, len(row_vals) + 1):
                        dv = ws.cell(row=r, column=c2).value
                        if dv not in (None, ""):
                            ledger_date = _to_date(dv)
                            break
                    break
        cell_val = row_vals[0] if row_vals else None
        if cell_val is None or str(cell_val).strip() == "" or _norm(cell_val) == "date":
            continue
        txt = str(cell_val).strip()
        if title is None:
            title = txt
        elif address_line is None:
            address_line = txt

    # 2) Read data rows after the header row.
    rows = []
    side_pairs = []  # (label, amount) from columns not part of the main col_map
    main_cols = set(col_map.keys())
    r = header_row_idx + 1
    while r <= ws.max_row:
        row_vals = {c: ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 15) + 1)}
        if all((v is None or str(v).strip() == "") for v in row_vals.values()):
            r += 1
            continue
        joined = " ".join(str(v) for v in row_vals.values() if v is not None).lower()
        if "total" in joined or "grand total" in joined:
            r += 1
            continue

        # Side table: cells in non-main columns forming (label, number) pairs.
        for c in range(1, min(ws.max_column, 15) + 1):
            if c in main_cols:
                continue
            val = row_vals.get(c)
            if isinstance(val, (int, float)) and val > 0:
                label_col = c - 1
                lbl = row_vals.get(label_col)
                if lbl and isinstance(lbl, str) and lbl.strip():
                    side_pairs.append((str(lbl).strip(), float(val)))

        rec = {"vendor": None, "department": None, "notes": None, "qty": None,
               "amount": None, "expense_date": ledger_date}

        for c, key in col_map.items():
            if key == "_sl":
                continue
            val = row_vals.get(c)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            if key == "qty":
                rec["qty"] = str(val).strip()
            elif key == "amount":
                try:
                    rec["amount"] = float(val)
                except (ValueError, TypeError):
                    rec["amount"] = None
            else:
                rec[key] = str(val).strip()

        # If amount cell empty, scan columns after qty for a stray numeric value
        # (handles the ledger's split sub-amount quirk).
        if rec["amount"] is None:
            qty_col = next((c for c, k in col_map.items() if k == "qty"), None)
            start = (qty_col + 1) if qty_col else 1
            for c in range(start, min(ws.max_column, 15) + 1):
                if c in col_map:
                    continue
                val = row_vals.get(c)
                if isinstance(val, (int, float)) and val > 0:
                    rec["amount"] = float(val)
                    break

        if rec["amount"] is not None or rec["notes"] or rec["vendor"]:
            rows.append(rec)
        r += 1

    # 3) Append side-table misc costs as additional expense rows.
    for label, amt in side_pairs:
        rows.append({
            "vendor": None, "department": None,
            "notes": label, "qty": None, "amount": amt,
            "expense_date": ledger_date,
        })

    if not title:
        title = "Imported Requisition"

    address = address_line or None
    period = None
    if address_line:
        import re
        m = re.search(
            r"(.*?)[\s,–\-]*((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*"
            r"(?:[\s\-–]+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*)?"
            r"[\s\-–]*\d{4}(?:[\s\-–/]+\d{4})?)\s*$",
            address_line, re.IGNORECASE,
        )
        if m:
            address = m.group(1).strip(" ,–-") or None
            period = m.group(2).strip()

    return title, address, period, ledger_date, rows


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Upload an Excel ledger (.xlsx) and auto-create a requisition with all rows."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload a .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read Excel file: {exc}")

    try:
        title, address, period, ledger_date, rows = _parse_import_workbook(wb)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Could not parse the Excel layout: {exc}")

    try:
        req = Requisition(
            id=str(uuid.uuid4()),
            title=title,
            status="open",
            address=address,
            period=period,
            ledger_date=ledger_date,
        )
        db.add(req)
        db.flush()

        created = 0
        for rec in rows:
            amt = rec.get("amount")
            if amt is None:
                continue  # skip rows without a usable amount
            exp = RequisitionExpense(
                id=str(uuid.uuid4()),
                requisition_id=req.id,
                expense_date=rec.get("expense_date") or date_type.today(),
                notes=rec.get("notes"),
                amount=amt,
                vendor=rec.get("vendor"),
                department=rec.get("department"),
                qty=rec.get("qty"),
            )
            db.add(exp)
            created += 1

        db.commit()
        db.refresh(req)
    except Exception as exc:
        db.rollback()
        # Surface the real error so the user sees *why* it failed instead of a
        # generic 500.  Common causes: missing DB table/column (needs a server
        # restart so the startup migrations run), or a schema mismatch.
        raise HTTPException(500, f"Failed to save requisition: {exc}")

    data = _req_to_dict(req)
    data["expenses"] = [_exp_to_dict(e) for e in req.expenses]
    return {
        "success": True,
        "data": data,
        "imported_rows": created,
        "message": f"Requisition created with {created} expense row(s)",
    }


@router.delete("/{req_id}")
def delete_requisition(
    req_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    req = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    db.delete(req)
    db.commit()
    return {"success": True, "message": "Requisition deleted"}
